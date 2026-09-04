"""
Разовый скрипт миграции: добавляет в существующий youtube.xlsx
листы "Статистика по месяцам" и "Статистика по годам", если их там ещё нет.

Запустить один раз из той же папки, где лежит youtube.xlsx:
    py migrate.py
"""

import os
from openpyxl import load_workbook

from main import FILE_PATH, DAY_SHEET, MONTH_SHEET, YEAR_SHEET, STATS_SHEETS, updateStatsSheets
from excelFormatter import applyFullStyle


def migrate():
    if not os.path.exists(FILE_PATH):
        print(f"❌ Файл не найден: {FILE_PATH}")
        return

    wb = load_workbook(FILE_PATH)

    added = []
    for name in STATS_SHEETS:
        if name not in wb.sheetnames:
            wb.create_sheet(name)
            added.append(name)

    if not added:
        print("Все листы уже на месте, миграция не требуется.")
        return

    print("Добавлены листы:", ", ".join(added))

    # пересчитываем статистику по всем листам (дни/месяцы/годы)
    updateStatsSheets(wb)

    # применяем оформление ко всем листам
    applyFullStyle(wb["Просмотры"])
    for name in STATS_SHEETS:
        applyFullStyle(wb[name])

    wb.save(FILE_PATH)
    print("✔ Готово. Файл обновлён:", FILE_PATH)


if __name__ == "__main__":
    migrate()
