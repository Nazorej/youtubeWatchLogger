from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# ---------- ШАПКА ----------

def styleHeader(ws):
    headerFill = PatternFill(
        start_color="C6EFCE",
        end_color="C6EFCE",
        fill_type="solid"
    )

    headerFont = Font(bold=True)

    headerAlignment = Alignment(
        horizontal="center",
        vertical="top",       # важно для wrap_text
        wrap_text=True
    )

    # немного увеличим высоту строки
    ws.row_dimensions[1].height = 30

    for cell in ws[1]:
        cell.fill = headerFill
        cell.font = headerFont
        cell.alignment = headerAlignment


# ---------- ГРАНИЦЫ ----------

def addBorders(ws):
    thin = Side(style="thin")

    border = Border(
        left=thin,
        right=thin,
        top=thin,
        bottom=thin
    )

    for row in ws.iter_rows():
        for cell in row:
            cell.border = border


# ---------- ВЫРАВНИВАНИЕ ДАННЫХ ----------

def alignContent(ws):
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="center")


# ---------- ЗЕБРА ----------

def zebraRows(ws):
    fill = PatternFill(
        start_color="F2F2F2",
        end_color="F2F2F2",
        fill_type="solid"
    )

    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        if i % 2 == 0:
            for cell in row:
                cell.fill = fill


# ---------- ШИРИНА КОЛОНОК ----------

def setColumnWidths(ws):
    if ws.title == "Просмотры":
        widths = {
            "A": 12,  # Дата
            "B": 10,  # Время
            "C": 28,  # Автор
            "D": 20,  # ID
            "E": 60   # Название
        }

    elif ws.title == "Статистика по дням":
        widths = {
            "A": 12,
            "B": 22  # чтобы точно влез "Просмотрено"
        }

    elif ws.title == "Статистика по месяцам":
        widths = {
            "A": 14,
            "B": 22
        }

    elif ws.title == "Статистика по годам":
        widths = {
            "A": 10,
            "B": 22
        }

    else:
        widths = {}

    for col, width in widths.items():
        ws.column_dimensions[col].width = width


# ---------- ФОРМАТЫ ДАТЫ / ВРЕМЕНИ ----------

# "[h]:mm:ss" — формат "длительности", а не времени суток: он корректно
# показывает и 5 часов, и 130 часов (когда сумма за месяц/год превышает сутки).
DURATION_FORMAT = "[h]:mm:ss"
DATE_FORMAT = "DD.MM.YYYY"
MONTH_FORMAT = "MM.YYYY"
YEAR_FORMAT = "0"


def applyDateTimeFormats(ws):
    if ws.title == "Просмотры":
        for row in ws.iter_rows(min_row=2):
            row[0].number_format = DATE_FORMAT      # Дата
            row[1].number_format = DURATION_FORMAT  # Время (длительность видео)

    elif ws.title == "Статистика по дням":
        for row in ws.iter_rows(min_row=2):
            row[0].number_format = DATE_FORMAT
            row[1].number_format = DURATION_FORMAT

    elif ws.title == "Статистика по месяцам":
        for row in ws.iter_rows(min_row=2):
            row[0].number_format = MONTH_FORMAT
            row[1].number_format = DURATION_FORMAT

    elif ws.title == "Статистика по годам":
        for row in ws.iter_rows(min_row=2):
            row[0].number_format = YEAR_FORMAT
            row[1].number_format = DURATION_FORMAT


# ---------- UX ----------

def freezeHeader(ws):
    ws.freeze_panes = "A2"


# ---------- ГЛАВНАЯ ФУНКЦИЯ ----------

def applyFullStyle(ws):
    applyDateTimeFormats(ws)
    styleHeader(ws)
    setColumnWidths(ws)
    addBorders(ws)
    alignContent(ws)
    zebraRows(ws)
    freezeHeader(ws)
