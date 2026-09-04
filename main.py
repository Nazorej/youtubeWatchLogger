import os
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook

from excelFormatter import applyFullStyle
from ytService import fetchVideoData


FILE_PATH = os.path.join(os.getcwd(), "youtube.xlsx")

DAY_SHEET = "Статистика по дням"
MONTH_SHEET = "Статистика по месяцам"
YEAR_SHEET = "Статистика по годам"
STATS_SHEETS = [DAY_SHEET, MONTH_SHEET, YEAR_SHEET]


def openOrCreateWorkbook():
    if os.path.exists(FILE_PATH):
        return load_workbook(FILE_PATH)

    wb = Workbook()

    ws = wb.active
    ws.title = "Просмотры"
    ws.append(["Дата", "Время", "Автор", "ID", "Название"])

    for name in STATS_SHEETS:
        wb.create_sheet(name)

    return wb


def isDuplicate(ws, videoId: str) -> bool:
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[3] == videoId:
            return True
    return False


def appendVideo(ws, data: dict):
    ws.append([
        data["date"],
        data["duration"],
        data["author"],
        data["video_id"],
        data["title"]
    ])


def _toDate(value):
    """Приводит значение из колонки 'Дата' к datetime.date.
    Поддерживает как новые записи (date-объект), так и старые (строка dd.mm.yy)."""
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, str):
        return datetime.strptime(value, "%d.%m.%y").date()
    return value  # уже datetime.date


def _toSeconds(value) -> int:
    """Приводит значение из колонки 'Время' к количеству секунд.
    Поддерживает как новые записи (timedelta), так и старые (строка HH:MM:SS)."""
    if isinstance(value, timedelta):
        return int(value.total_seconds())
    if isinstance(value, str):
        h, m, s = map(int, value.split(":"))
        return h * 3600 + m * 60 + s
    if hasattr(value, "hour"):  # datetime.time (на случай старого формата)
        return value.hour * 3600 + value.minute * 60 + value.second
    return 0


def _rebuildStatsSheet(statsWs, headerLabel: str, stats: dict):
    statsWs.delete_rows(1, statsWs.max_row)
    statsWs.append([headerLabel, "Просмотрено"])

    for key in sorted(stats.keys()):
        statsWs.append([key, timedelta(seconds=stats[key])])


def updateStatsSheets(wb):
    ws = wb["Просмотры"]

    dayStats = {}
    monthStats = {}
    yearStats = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        date, duration = row[0], row[1]

        if not date or duration is None:
            continue

        dateObj = _toDate(date)
        seconds = _toSeconds(duration)

        dayStats[dateObj] = dayStats.get(dateObj, 0) + seconds

        monthKey = dateObj.replace(day=1)
        monthStats[monthKey] = monthStats.get(monthKey, 0) + seconds

        yearKey = dateObj.year
        yearStats[yearKey] = yearStats.get(yearKey, 0) + seconds

    _rebuildStatsSheet(wb[DAY_SHEET], "Дата", dayStats)
    _rebuildStatsSheet(wb[MONTH_SHEET], "Месяц", monthStats)
    _rebuildStatsSheet(wb[YEAR_SHEET], "Год", yearStats)


def processUrl(url: str):
    data = fetchVideoData(url)
    if not data:
        print("❌ Не удалось получить данные")
        return

    wb = openOrCreateWorkbook()
    ws = wb["Просмотры"]

    if isDuplicate(ws, data["video_id"]):
        print("⚠ Уже есть:", data["title"])
        return

    appendVideo(ws, data)
    updateStatsSheets(wb)

    # применяем стиль ко всем листам
    applyFullStyle(ws)
    for name in STATS_SHEETS:
        applyFullStyle(wb[name])

    wb.save(FILE_PATH)

    print("✔ Добавлено:", data["title"])


if __name__ == "__main__":
    processUrl("https://youtu.be/ВАШ_ID")
